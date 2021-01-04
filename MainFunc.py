
import openpyxl
import pyodbc
import time

conn_String = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\6408284\Desktop\새 폴더 (2)\moduleBook.accdb;'
file_Name = r"C:\Users\6408284\Desktop\원단위 2 - 12.10 이준우.xlsx"
# file_Name = r"C:\Users\6408284\Desktop\새 폴더 (2)\청구내역2011-2019.xlsx"

output_file_Name = r"C:\Users\6408284\Desktop\원단위 2 - 12.10 이준우.xlsx"
sht_List = ["BR_YB", "BR_NE", "BR_OE", "BR_OS EV", "BR_DL3", "BR_JK EV", "BR_DL3c", "BR_DN8c", "BR_TMc"]
sht_Ref_List = ["JX1", "IK", "MQ4", "BD", "TL", "DN8", "DH"]
sht_china_List = ["BR_DL3c", "BR_DN8c", "BR_TMc"]
rd_column_index_dict = {"TL": 4, "JX1": 7, "DN8": 10, "MQ4": 13, "IK": 16, "BD": 19, "DH": 22}

# 시작의장BD/ 시작의장TL,DN8,DH / 시작의장MQ4/ 시작의장IK/ 시작의장JX(전산)
row_index = 9
# 모듈명입력위치
col_index = 9
# UPG 위치
upg_location = 6
base_ratio = 0.65

def distribution(conn_String,file_Name, sht_Name,output_file_Name,row_index,col_index,upg_location):

    targetBk = openpyxl.load_workbook(file_Name)
    targetSht = targetBk[sht_Name]

    searchUPG = ""

    for i in range(row_index, targetSht.max_row + 1):

        originUPG = str(targetSht.cell(row=i, column=upg_location).value)
        searchUPG = upg_replace(originUPG)

        #print(searchUPG)

        selectQuery = "select * from ModuleBOM_eBOM where eBOMUPG ='" + searchUPG + "'"
        conn = pyodbc.connect(conn_String)
        cursor = conn.cursor()
        cursor.execute(selectQuery)
        q_list = list(cursor.fetchall())
        cursorLength = len(q_list)
        #cursor.rollback()
        #cursor.execute(selectQuery)
        # for row in cursor.fetchall():

        if cursorLength == 1:
            for row in q_list:
                # print(row[3])
                selectQuery2 = "select * from ModuleList where ID ='" + str(row[3])+"'"
                cursor2 = conn.cursor()
                cursor2.execute(selectQuery2)
                for row2 in cursor2.fetchall():
                    targetSht.cell(row=i, column=col_index).value = row2[3]
                    #targetSht.cell(row=i, column=8).value = row2[4]
        elif cursorLength > 1:
            print("ERROR cursorLength : " + str(cursorLength) + ":" + searchUPG)

        elif cursorLength == 0 and searchUPG != "":
            selectQuery3 = "select UPG_NAME from NonModuleList where UPG_NO ='" + searchUPG + "'"
            cursor3 = conn.cursor()
            cursor3.execute(selectQuery3)
            if len(cursor3.fetchall()) >= 1:
                targetSht.cell(row=i, column=col_index).value = "비모듈"

    targetBk.save(output_file_Name)
    targetBk.close()

def material_match(conn_String,file_Name, sht_Name,output_file_Name,row_index,col_index,upg_location):

    targetBk = openpyxl.load_workbook(file_Name)
    targetSht = targetBk[sht_Name]
    print(sht_Name)
    searchUPG = ""

    for i in range(row_index, targetSht.max_row + 1):


        originUPG = str(targetSht.cell(row=i, column=upg_location).value)
        searchUPG = upg_replace(originUPG)

        print(searchUPG)

        selectQuery = "select * from ModuleBOM_eBOM where eBOMUPG ='" + searchUPG + "'"
        conn = pyodbc.connect(conn_String)
        cursor = conn.cursor()
        cursor.execute(selectQuery)
        cursorLength = len(cursor.fetchall())
        cursor.rollback()
        cursor.execute(selectQuery)
        # for row in cursor.fetchall():

        if cursorLength == 1:
            for row in cursor.fetchall():
                print(row[3])
                selectQuery2 = "select * from ModuleList where ID ='" + str(row[3])+"'"
                cursor2 = conn.cursor()
                cursor2.execute(selectQuery2)
                for row2 in cursor2.fetchall():
                    targetSht.cell(row=i, column=col_index).value = row2[3]
                    targetSht.cell(row=i, column=col_index-7).value = row2[1]
                    targetSht.cell(row=i, column=col_index-6).value = row2[2]
                    if row2[4] == "Strategic":
                        targetSht.cell(row=i, column=col_index+1).value = "●"
                    elif row2[4] == "Universal":
                        targetSht.cell(row=i, column=col_index+2).value = "●"
                    elif row2[4] == "Local":
                        targetSht.cell(row=i, column=col_index+3).value = "●"
                    elif row2[4] == "Local(Design)":
                        targetSht.cell(row=i, column=col_index+4).value = "●"

        elif cursorLength > 1:
            print("ERROR cursorLength : " + str(cursorLength) + ":" + searchUPG)

        elif cursorLength == 0 and searchUPG != "":
            selectQuery3 = "select UPG_NAME from NonModuleList where UPG_NO ='" + searchUPG + "'"
            cursor3 = conn.cursor()
            cursor3.execute(selectQuery3)
            if len(cursor3.fetchall()) >= 1:
                targetSht.cell(row=i, column=col_index).value = "비모듈"

    targetBk.save(output_file_Name)
    targetBk.close()

def upg_replace(originUPG):

    if len(originUPG) == 6:
        originUPG = originUPG + "E"

    searchUPG = originUPG
    if originUPG != None:
        if str.find(originUPG, "/") != -1:
            searchUPG = originUPG[:7]
            # print(searchUPG+"/ TPYE")
        elif len(originUPG) == 12 and str.find(originUPG, "UPG") != -1:
            searchUPG = originUPG[-7:]
            # print(searchUPG+"include UPG")
        elif len(originUPG) == 12 and str.find(originUPG, "UPG") == -1:
            searchUPG = originUPG[:7]
            # print(searchUPG+"include VC")
        elif len(originUPG) == 11 and str.find(originUPG, "UPG") != -1:
            searchUPG = originUPG[-6:] + "E"
            # print(searchUPG)

    else:
        searchUPG = ""

    return searchUPG

def input_vehicleCost(targetSht):
    #Lampup data
    localBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_local.xlsm", read_only=False, keep_vba=True)
    nonlocalBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_universal_strategic.xlsm", read_only=False, keep_vba=True)
    nonlocalBk_base = openpyxl.load_workbook(
        filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_universal_strategic_mKRW.xlsm", read_only=False,
        keep_vba=True)

    localDataSht = localBk["Data"]
    nonlocalDataSht = nonlocalBk["Data"]
    nonlocalData_baseSht = nonlocalBk_base["Data"]

    # 데이터 초기화
    localDataSht.delete_rows(2, localDataSht.max_row+1)
    nonlocalDataSht.delete_rows(2, nonlocalDataSht.max_row + 1)
    nonlocalData_baseSht.delete_rows(2,  nonlocalData_baseSht.max_row + 1)

    car_code = ""
    seq_nonLocal = 2
    seq_local = 2

    direct_nonlocal_labor_value = 0
    direct_local_labor_value = 0
    direct_nonlocal_other_value = 0
    direct_local_other_value = 0

    for car,i in rd_column_index_dict.items() :


        selectQuery = "select * from VehicleData where Car_Name ='" + car + "'"
        conn = pyodbc.connect(conn_String)
        cursor = conn.cursor()
        cursor.execute(selectQuery)
        # print(car)
        # print(i)
        for row in cursor.fetchall():
            car_code = row[0]

        for j in range(3, 338):
            module_code = targetSht.cell(row=j, column=2).value
            selectQuery2 = "select * from ModuleList where ID ='" + module_code + "'"
            cursor2 = conn.cursor()
            cursor2.execute(selectQuery2)
            for row in cursor2.fetchall():
                if (row[4] == "Strategic" or row[4] == "Universal") and module_code !="MO0267" and module_code !="MO0333"and module_code !="MO0334":
                    direct_nonlocal_labor_value = targetSht.cell(row=j, column=i+1).value
                    direct_nonlocal_other_value = targetSht.cell(row=j, column=i).value

                    if direct_nonlocal_other_value!= 0 and direct_nonlocal_other_value is not None:
                        nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal, car_code,module_code,"NewPAI030", direct_nonlocal_other_value*(1-base_ratio))
                        nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal, car_code, module_code,"NewPAI042", direct_nonlocal_other_value * base_ratio)
                        seq_nonLocal = seq_nonLocal + 1
                    if direct_nonlocal_labor_value != 0 and direct_nonlocal_labor_value is not None:
                        nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal, car_code, module_code, "NewPAI029", direct_nonlocal_labor_value*(1-base_ratio))
                        nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal, car_code, module_code,"NewPAI041", direct_nonlocal_labor_value * base_ratio)
                        seq_nonLocal = seq_nonLocal + 1

                elif module_code =="MO0267":

                    nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal, car_code, module_code, "NewPAI030",
                                                    targetSht.cell(row=j, column=i).value)
                    nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal+1, car_code, module_code, "NewPAI029",
                                                    targetSht.cell(row=j, column=i+1).value)

                    nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal, car_code, module_code,
                                                         "NewPAI042", targetSht.cell(row=366, column=i).value)
                    nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal+1, car_code, module_code, "NewPAI041",
                                                        targetSht.cell(row=366, column=i+1).value)

                    seq_nonLocal = seq_nonLocal + 2

                elif module_code =="MO0333":

                    nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal, car_code, module_code, "NewPAI030",
                                                    targetSht.cell(row=j, column=i).value)
                    nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal + 1, car_code, module_code,
                                                    "NewPAI029",
                                                    targetSht.cell(row=j, column=i + 1).value)

                    nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal, car_code, module_code,
                                                         "NewPAI042", targetSht.cell(row=367, column=i).value)
                    nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal + 1, car_code, module_code,
                                                         "NewPAI041",
                                                         targetSht.cell(row=367, column=i+1).value)

                    seq_nonLocal = seq_nonLocal + 2

                elif module_code =="MO0334":

                    nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal, car_code, module_code, "NewPAI030",
                                                    targetSht.cell(row=j, column=i).value)
                    nonlocalDataSht = rd_cell_input(nonlocalDataSht, seq_nonLocal + 1, car_code, module_code,
                                                    "NewPAI029",
                                                    targetSht.cell(row=j, column=i + 1).value)

                    nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal, car_code, module_code,
                                                         "NewPAI042", targetSht.cell(row=368, column=i).value)
                    nonlocalData_baseSht = rd_cell_input(nonlocalData_baseSht, seq_nonLocal + 1, car_code, module_code,
                                                         "NewPAI041",
                                                         targetSht.cell(row=368, column=i+1).value)

                    seq_nonLocal = seq_nonLocal + 2

                elif row[4] == "Local" or row[4] == "Local(Design)":

                    direct_local_labor_value = targetSht.cell(row=j, column=i+1).value
                    direct_local_other_value = targetSht.cell(row=j, column=i).value
                    if direct_local_other_value != 0:
                        localDataSht = rd_cell_input(localDataSht, seq_local, car_code, module_code, "NewPAI028", direct_local_other_value)
                        seq_local = seq_local + 1
                    if direct_local_labor_value != 0:
                        localDataSht = rd_cell_input(localDataSht, seq_local, car_code, module_code, "NewPAI026", direct_local_labor_value)
                        seq_local = seq_local + 1

    localBk .save(r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_local.xlsm")
    localBk.close()

    nonlocalBk.save(r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_universal_strategic.xlsm")
    nonlocalBk.close()

    nonlocalBk_base.save(r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_universal_strategic_mKRW.xlsm")
    nonlocalBk_base.close()


def input_lamp_up_indirect_vehicle(targetSht):

    # Lampup data
    LampupBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\Ramp-up_Cost.xlsm", read_only=False, keep_vba=True)
    indirectBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_indirect.xlsm",
                                      read_only=False, keep_vba=True)
    vehicleBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_vehicle.xlsm",
                                      read_only=False, keep_vba=True)
    lampupDataSht = LampupBk["Data"]
    indirectDataSht = indirectBk ["Data"]
    vehicleDataSht = vehicleBk["Data"]

    # 데이터 초기화
    lampupDataSht.delete_rows(2,lampupDataSht.max_row+1)
    indirectDataSht.delete_rows(2, indirectDataSht.max_row + 1)
    vehicleDataSht.delete_rows(2,  vehicleDataSht.max_row + 1)
    seq_lumpup_vehicle= 2
    seq_indirect = 2

    car_code =""
    eo_value =0
    pilot_value = 0

    for car,i in rd_column_index_dict.items():

        selectQuery = "select * from VehicleData where Car_Name ='" + car + "'"
        conn = pyodbc.connect(conn_String)
        cursor = conn.cursor()
        cursor.execute(selectQuery)
        print(car)
        #print(i)
        for row in cursor.fetchall():
            car_code = row[0]

        eo_value = targetSht.cell(row=361, column=i).value
        pilot_value = targetSht.cell(row=360, column=i).value

        vehicle_other_value = row_sum(targetSht,338,22,i) + targetSht.cell(row=369, column=i).value + targetSht.cell(row=370, column=i).value

        vehicle_labor_value =  targetSht.cell(row=365, column=i).value
        indirect_value = row_sum(targetSht,362,3,i)
        #print(eo_value)
        #print(pilot_value)

        lampupDataSht = rd_cell_input(lampupDataSht, seq_lumpup_vehicle,car_code,"dummy","NewPAI048",eo_value)
        lampupDataSht = rd_cell_input(lampupDataSht, seq_lumpup_vehicle+1, car_code, "dummy", "NewPAI037", pilot_value)

        vehicleDataSht = rd_cell_input(vehicleDataSht, seq_lumpup_vehicle,car_code,"dummy","NewPAI027",vehicle_other_value)
        vehicleDataSht = rd_cell_input(vehicleDataSht, seq_lumpup_vehicle+1, car_code, "dummy", "NewPAI025", vehicle_labor_value)

        indirectDataSht = rd_cell_input(indirectDataSht, seq_indirect,car_code,"dummy","NewPAI043",indirect_value)

        seq_lumpup_vehicle = seq_lumpup_vehicle + 2
        seq_indirect = seq_indirect + 1

    LampupBk.save(r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\Ramp-up_Cost.xlsm")
    indirectBk.save(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_indirect.xlsm")
    vehicleBk.save(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TopDown\R_D_vehicle.xlsm")

    LampupBk.close()
    indirectBk.close()
    vehicleBk.close()


def rd_cell_input(shtname, seq,car_code,module_code,pai_code,value):

    shtname.cell(row=seq, column=1).value = "Current"
    shtname.cell(row=seq, column=2).value = car_code
    shtname.cell(row=seq, column=3).value = "mKRW"
    shtname.cell(row=seq, column=4).value = module_code
    shtname.cell(row=seq, column=5).value = pai_code
    shtname.cell(row=seq, column=6).value = value
    shtname.cell(row=seq, column=7).value = "Current_" + car_code + "_mKRW" + "_" + module_code +"_"+pai_code

    return shtname

def input_ref_data() :

    targetBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\(201211) 신마스터 (335개)_종합본(인건비 통합)_result.xlsx",  data_only=True)
    targetSht = targetBk["종합"]

    input_lamp_up_indirect_vehicle(targetSht)
    input_vehicleCost(targetSht)


    targetBk.close()

def upgModuleTable():

    targetBk = openpyxl.load_workbook(r"C:\Users\6408284\Desktop\새 폴더 (2)\ARCHITECTURE.xlsx")
    dSht = targetBk["EBOM"]

    searchUPG = ""

    for i in range(2, dSht.max_row + 1):

        originUPG = str(dSht.cell(row=i, column=1).value)

        selectQuery = "select * from ModuleBOM_eBOM where eBOMUPG ='" + originUPG + "'"
        conn = pyodbc.connect(conn_String)
        cursor = conn.cursor()
        cursor.execute(selectQuery)
        q_list = list(cursor.fetchall())
        cursorLength = len(q_list)
        # cursor.rollback()
        # cursor.execute(selectQuery)
        # for row in cursor.fetchall():

        if cursorLength == 1:
            for row in q_list:
                #print(row[3])
                selectQuery2 = "select * from ModuleList where ID ='" + str(row[3])+ "'"
                dSht.cell(row=i, column=3).value = row[3]
                cursor2 = conn.cursor()
                cursor2.execute(selectQuery2)
                for row2 in cursor2.fetchall():
                    dSht.cell(row=i, column=4).value = row2[3]

        elif cursorLength > 1:
            print("ERROR cursorLength : " + str(cursorLength) + ":" + searchUPG)


    targetBk.save(r"C:\Users\6408284\Desktop\새 폴더 (2)\ARCHITECTURE.xlsx")
    targetBk.close()

def dragUPG():
    targetBk = openpyxl.load_workbook(r"C:\Users\6408284\Desktop\새 폴더 (2)\구매투자비_201125.xlsx")
    targetSht = targetBk["DH"]
    dragSht = targetBk["MQ4"]

    for i in range(7, targetSht.max_row + 1):
        for j in range(7, dragSht.max_row + 1):
            if targetSht.cell(row = i,column=10)  != None  and targetSht.cell(row = i,column=11)  != None and len(str(targetSht.cell(row = i,column=9).value))  != 7:
                if str(dragSht.cell(row = j,column=10).value)[:5] == str(targetSht.cell(row = i,column=10).value)[:5]:
                    targetSht.cell(row = i,column=9).value= dragSht.cell(row=j, column=9).value
                    break
                #elif str.find(str.replace(str(targetSht.cell(row = i,column=11).value)," ",""), str.replace(str(targetSht.cell(row = j,column=11).value)," ","")) != -1 :
                #    targetSht.cell(row=i, column=9).value = dragSht.cell(row=j, column=9).value
                #    break

    targetBk.save(r"C:\Users\6408284\Desktop\새 폴더 (2)\구매투자비_201125.xlsx")
    targetBk.close()

def proto_tcar_data_copy():
    print("Proto data input Start")
    start = time.time()
    protoBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\시작7개차종 DATA(335개 기준)_201211.xlsx",  data_only=True)
    tcarBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\새 폴더 (2)\TCAR.xlsx",  data_only=True)
    masterBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\(201211) 신마스터 (335개)_종합본(인건비 통합)_result.xlsx")
    protoSht = protoBk["종합"]
    tcarSht= tcarBk["종합"]
    seq = 0
    sumvalue = 0.0
    #masterSht = masterBk["종합"]
    print("LOAD time : " + str(time.time()-start))
    sht_time = time.time()

    for car_name in sht_Ref_List:

        print("SHT: " +car_name +"/ TIME: "+str(time.time()-sht_time))
        masterSht = masterBk[car_name]
        #proto data input(모듈)
        for i in range(3,365):
            sumvalue = protoSht.cell(row=i, column=2+seq*5).value
            if sumvalue !=0.0:
                for j in range(1, 4):
                    masterSht.cell(row=i + 3, column=133+j).value=protoSht.cell(row=i, column=2+seq*5+j).value
        #비모듈 데이터
        masterSht.cell(row=412, column=134).value = protoSht.cell(row=365, column=3 + seq * 5).value
        masterSht.cell(row=413, column=135).value = protoSht.cell(row=365, column=4 + seq * 5).value

        #tcar data input
        for i in range(3,365):
            sumvalue = tcarSht.cell(row=i, column=2+seq*3).value
            if sumvalue !=0.0:
                for j in range(1,3):
                    masterSht.cell(row=i + 3, column=129+j).value=tcarSht.cell(row=i, column=2+seq*3+j).value
        #비모듈 데이터
        masterSht.cell(row=412, column=130).value = tcarSht.cell(row=365, column=3 + seq * 3).value
        masterSht.cell(row=413, column=131).value = tcarSht.cell(row=365, column=4 + seq * 3).value

        seq=seq+1
        sht_time = time.time()

    masterBk.save(filename=r"C:\Users\6408284\Desktop\(201211) 신마스터 (335개)_종합본(인건비 통합)_result .xlsx")
    masterBk.close()
    protoBk.close()
    tcarBk.close()

def row_sum(sht_name,initial_row,num_of_cell,fixed_col) :
    sum = 0.0
    for i in range(0,num_of_cell):
        #print(str("value : "+ str(sht_name.cell(row=initial_row + i, column=fixed_col).value) +" " + str(i)+"  "+str(fixed_col)))
        if sht_name.cell(row=initial_row + i, column=fixed_col).value != None :
            sum = sum + float(sht_name.cell(row=initial_row + i, column=fixed_col).value)

    return sum


def gathering_ref_data():
    targetBk =openpyxl.load_workbook(r"C:\Users\6408284\Desktop\검증시트.xlsx")
    targetSht= targetBk["Sheet1"]
    targetSht.delete_rows(2,targetSht.max_row+1)

    root_path = "C:\\Users\\6408284\\Desktop\\새 폴더 (2)\\TopDown\\"
    seq=2

    workBkforRef=["Ramp-up_Cost","R_D_indirect","R_D_vehicle","R_D_local","R_D_universal_strategic","R_D_universal_strategic_mKRW"]

    for filenameRef in workBkforRef:
        filepath=str(root_path)+(filenameRef)+".xlsm"
        bkName = openpyxl.load_workbook(filename=filepath, read_only=False, keep_vba=True)
        dataSht =bkName["Data"]
        for i in range (2,dataSht.max_row+1):
            for j in range(1,dataSht.max_column+1 ):
                targetSht.cell(row=seq,column=j).value=dataSht.cell(row=i,column=j).value
            seq = seq + 1
        bkName.close()

    targetBk.save(r"C:\Users\6408284\Desktop\검증시트.xlsx")
    targetBk.close()

def replaceBTU():
    dataBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\BottomUp_old_RnD_Local_EV.xlsx",  data_only=True)
    dataSht = dataBk["Data_old"]
    targetSht = dataBk["Data_new"]
    transferSht = dataBk["OldToNew"]
    modulename = ""
    modename =""
    new_modulename =""
    split_moudlename =""
    seq=2

 #   pailist = ["NewPAI010","NewPAI013","NewPAI016","NewPAI019","NewPAI026","NewPAI028","NewPAI029","NewPAI030","NewPAI032",
  #             "NewPAI033","NewPAI034","NewPAI041","NewPAI042","NewPAI044","NewPAI045"]
    conn = pyodbc.connect(conn_String)
    cursor = conn.cursor()
    cursorforinsert = conn.cursor()

    for i in range(2,transferSht.max_row+1):


        modulename = transferSht.cell(row=i,column=1).value
        modename = transferSht.cell(row=i,column=3).value


        if modename =="SPLIT" :
            new_modulename = transferSht.cell(row=i, column=4).value
            split_modulename = transferSht.cell(row=i, column=5).value

        elif modename =="ADD-A" or modename =="ADD-B":
            new_modulename = transferSht.cell(row=i,column=4).value
        else :
            new_modulename = modename
            modename = "SAME"

        print("PHASE : "+ str(i)+" MODE NAME : "+modename)

        for carindex in range(1,22):

            carname = "VE00"

            if carindex <10:
                carname = carname + "0" + str(carindex)
            else:
                carname = carname + str(carindex)

            selectQuery = "select * from Data_old where Module ='" + modulename + "' and ID='"+carname+"'"
            cursor.execute(selectQuery)
            q_list = list(cursor.fetchall())

            cursorLength = len(q_list)

            if cursorLength >= 1 :
                if modename == "SAME":
                    for row in q_list:
                        insertQuery = "insert into Data_new values ('"+ carname+"','"+new_modulename+"','"+row[2]+"',"+str(row[3])+",'"+ new_modulename+"_"+row[2]+"_"+row[0]+"')"
                        #print("SAME:"+insertQuery)
                        cursorforinsert.execute(insertQuery)
                        conn.commit()

                elif modename == "ADD-A":
                    for row in q_list:
                        selectQuery2 = "select Value_mKRW from Data_new where index_ID ='"+new_modulename+"_"+row[2]+"_"+row[0]+"'"
                        print("ADD-A:"+selectQuery2)
                        cursor.execute(selectQuery2)
                        c_list = list(cursor.fetchall())
                        c_list_length = len(c_list)
                        old_value = 0
                        if c_list_length >= 1:

                            for row2 in c_list :
                                old_value = row2[0]

                            updateQuery ="UPDATE Data_new SET Value_mKRW='"+str(old_value+row[3])+"' WHERE Index_ID ='"+new_modulename+"_"+row[2]+"_"+row[0]+"'"
                            print(updateQuery)
                            cursorforinsert.execute(updateQuery)
                            conn.commit()

                        if c_list_length < 1:
                            insertQuery = "insert into Data_new values ('" + carname + "','" + new_modulename + "','" + row[
                                2] + "'," + str(row[3]) + ",'" + new_modulename + "_" + row[2] + "_" + row[0] + "')"
                            print("ADD-A insert:"+insertQuery)
                            cursorforinsert.execute(insertQuery)
                            conn.commit()

                elif modename == "SPLIT" :
                    for row in q_list:
                        insertQuery = "insert into Data_new values ('" + carname + "','" + new_modulename + "','" + row[
                            2] + "'," + str(row[3]/2) + ",'" + new_modulename + "_" + row[2] + "_" + row[0] + "')"
                        #print("SPLIT-1:"+insertQuery)
                        cursorforinsert.execute(insertQuery)
                        conn.commit()

                        insertQuery = "insert into Data_new values ('" + carname + "','" + split_modulename + "','" + row[
                            2] + "'," + str(row[3]/2) + ",'" + split_modulename + "_" + row[2] + "_" + row[0] + "')"
                        #print("SPLIT-2:"+insertQuery)
                        cursorforinsert.execute(insertQuery)
                        conn.commit()

            else :
                if modename == "ADD-B":

                    selectQuery3 = "select BottomUp1, PAI from BaseData_old where eVehicle_ID ='"+carname+"' and Module_ID ='"+modulename+"'"
                    #print("ADD-B for base data:"+selectQuery3)
                    cursor.execute(selectQuery3)
                    b_list = list(cursor.fetchall())



                    for row3 in b_list :
                        selectQuery2 = "select Value_mKRW from Data_new where index_ID ='" + new_modulename + "_" + row3[
                             1] + "_" + carname + "'"
                        print("ADD-B:" + selectQuery2)
                        cursor.execute(selectQuery2)
                        c_list = list(cursor.fetchall())
                        c_list_length = len(c_list)
                        old_value = 0
                        volume=1

                        if row3[1]=="NewPAI010" or row3[1]=="NewPAI013" or row3[1]=="NewPAI016" or row3[1]=="NewPAI019" or row3[1]=="NewPAI022" :
                            selectQuery4 = "select sum(Volume) from Volume_old where ID='" + carname + "'"
                            cursor.execute(selectQuery4)
                            volume_list = list(cursor.fetchall())
                            volume=volume_list[0][0]*0.001

                        if c_list_length >=1 :
                            for row2 in c_list:
                                old_value = row2[0]

                            updateQuery = "update data_new set Value_mKRW ='" + str(
                                    old_value + row3[0]/volume) + "' where index_ID ='" + new_modulename + "_" + row3[1] + "_" + carname + "'"
                            cursorforinsert.execute(updateQuery)
                            conn.commit()
                        else:
                            insertQuery = "insert into Data_new values ('" + carname + "','" + new_modulename + "','" + row[
                                2] + "'," + str(row3[0]/volume) + ",'" + new_modulename + "_" + row3[1] + "_" + carname + "')"
                            print("ADD-A insert:"+insertQuery)
                            cursorforinsert.execute(insertQuery)
                            conn.commit()

    dataBk.close()

def replaceFUC():

    dataBk = openpyxl.load_workbook(filename=r"C:\Users\6408284\Desktop\Func_change.xlsx",  data_only=True)
    transferSht = dataBk["Mo_chg_table"]
    conn = pyodbc.connect(conn_String)
    cursor = conn.cursor()

    module_name = ""
    mode_name =""
    new_module_name =""
    split_module_name =""

    for i in range(2,transferSht.max_row+1):

        module_name = transferSht.cell(row = i, column = 1).value
        mode_name = transferSht.cell(row = i, column = 5).value
        new_module_name = transferSht.cell(row=i, column=6).value

        if transferSht.cell(row = i, column = 4).value == "○":

            result_list = select_query_list(cursor,select_data_query_module("Data_old_Func",module_name,"*"))

            for q_list in result_list:

                if mode_name == "SAME":

                    commit_query_exe(conn,cursor,insert_data_query("Data_new_Func", q_list[0],new_module_name, q_list[2], str(q_list[3])))

                elif mode_name == "SPLIT":

                    split_module_name = transferSht.cell(row=i, column=7).value
                    commit_query_exe(conn, cursor,
                                     insert_data_query("Data_new_Func", q_list[0], new_module_name, q_list[2],
                                                       str(q_list[3]/2)))
                    commit_query_exe(conn, cursor,
                                     insert_data_query("Data_new_Func", q_list[0],  split_module_name, q_list[2],
                                                       str(q_list[3]/2)))

                elif mode_name == "ADD":

                    add_list = select_query_list(cursor,select_data_query_index("Data_new_Func", q_list[0], new_module_name, q_list[2],"*"))
                    if len(add_list) == 1:
                        commit_query_exe(conn, cursor,
                                        update_data_query("Data_new_Func", q_list[0], new_module_name, q_list[2], str(q_list[3]+add_list[0][3])))
                    else:
                        commit_query_exe(conn, cursor,
                                        insert_data_query("Data_new_Func", q_list[0], new_module_name, q_list[2], str(q_list[3])))

        elif transferSht.cell(row = i, column = 4).value == "△":

            if mode_name == "ADD":

                result_list = select_query_list(cursor, select_base_data_query("BaseData_old", module_name,
                                                                     "eVehicle_ID,Module_ID,PAI,FctChangesValue_mKRW"))
                for q_list in result_list:

                    if q_list[3] is None: q_list[3] = 0

                    add_list = select_query_list(cursor,select_data_query_index("Data_new_Func", q_list[0], new_module_name, q_list[2],"*"))
                    volume = 1

                    if q_list[2] == "NewPAI010" or q_list[2] == "NewPAI013" or q_list[2] == "NewPAI016" or q_list[2] == "NewPAI019" or q_list[2] == "NewPAI022":
                        selectQuery4 = "select sum(Volume) from Volume_old where ID='" + q_list[0] + "'"
                        cursor.execute(selectQuery4)
                        volume_list = list(cursor.fetchall())
                        volume = volume_list[0][0] * 0.001

                    if len(add_list) == 1:

                        commit_query_exe(conn, cursor,
                                         update_data_query("Data_new_Func", q_list[0], new_module_name, q_list[2],
                                                           str(q_list[3] + add_list[0][3]/volume)))
                    else:
                        commit_query_exe(conn, cursor,
                                         insert_data_query("Data_new_Func", q_list[0], new_module_name, q_list[2],
                                                           str(q_list[3]/volume)))


def make_car_name(number_of_car):

    car_name_list =[]
    for i in range(1, number_of_car):

        if i < 10:
            car_name = "VE000" + str(i)
            car_name_list.append(car_name)

        elif 10 <= i < 100:
            car_name = "VE00" + str(i)
            car_name_list.append(car_name)

        elif 100 <= i < 1000:
            car_name = "VE0" + str(i)
            car_name_list.append(car_name)

    return car_name_list


def insert_data_query(table_name, car_name, module_name, pai, value):
    insertQuery = "insert into "+table_name+" values ('"+car_name+"','"+module_name+"','"+pai+"','"+value+"','"+module_name+"_"+pai+"_"+car_name+"')"
    print(insertQuery)
    return insertQuery


def update_data_query(table_name, car_name, module_name, pai, value):
    updateQuery = "update " + table_name + " set Value_mKRW = '" + value + "' where Index_ID='"+module_name+"_"+pai+"_"+car_name+"'"
    print(updateQuery)
    return updateQuery


def select_data_query_index(table_name, car_name, module_name, pai, value):
    selectQuery = "select "+ value+" from "+table_name +" where Index_ID='"+module_name+"_"+pai+"_"+car_name+"'"
    print(selectQuery)
    return selectQuery


def select_data_query_module(table_name, module_name, value):
    selectQuery = "select "+ value+" from "+table_name +" where Module='"+module_name+"'"
    print(selectQuery)
    return selectQuery


def select_base_data_query(table_name, module_name, value):
    selectQuery = "select "+ value+" from "+table_name +" where Module_ID ='"+module_name+"'"
    print(selectQuery)
    return selectQuery


def select_query_list(cursor, select_query):
    cursor.execute(select_query)
    return list(cursor.fetchall())


def commit_query_exe(conn, cursor, commit_query):
    cursor.execute(commit_query)
    conn.commit()



# 엑셀파일 열기

# distribution(conn_String,file_Name, sht_Name,output_file_Name,row_index,col_index,upg_location)

#material_match(conn_String,file_Name, sht_Name,output_file_Name,row_index,col_index,upg_location)
#for sht_Names in sht_china_List:
#    material_match(conn_String, file_Name, sht_Names, output_file_Name, row_index, col_index, upg_location)
#material_match(conn_String, file_Name, "BR_TMc", output_file_Name, row_index, col_index, upg_location)
# dragUPG()
# upgModuleTable()


# input_ref_data()
# gathering_ref_data()

# Proto data input code
# proto_tcar_data_copy()

#replaceBTU()
replaceFUC()