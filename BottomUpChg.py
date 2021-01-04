import openpyxl
import pyodbc

file_Name = r"C:\Users\6408284\Desktop\★(201230)Bottom up 마스터(7.7용).xlsx"
data_File_Name = r"C:\Users\6408284\Desktop\Bottom_Up_RnD.xlsm"
conn_String = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\6408284\Desktop\새 폴더 (2)\moduleBook.accdb;'
base_rate = 0.65
labor_rate = 0.929

car_List = ["VE0012", "VE0014", "VE0015", "VE0017", "VE0013", "VE0011", "VE0006", "VE0002",
            "VE0003", "VE0007", "VE0016", "VE0010", "VE0018", "VE0021", "VE0008", "VE0004",
            "VE0005", "VE0009", "VE0001", "VE0019", "VE0020"]

china_car_List = ["VE0016", "VE0010", "VE0018", "VE0021", "VE0019", "VE0020"]

def cell_input(shtname, seq,car_code,module_code,pai_code,value):

    shtname.cell(row=seq, column=1).value = car_code
    shtname.cell(row=seq, column=2).value = module_code
    shtname.cell(row=seq, column=3).value = pai_code
    shtname.cell(row=seq, column=4).value = value
    shtname.cell(row=seq, column=5).value = module_code + "_" + pai_code +"_"+car_code

    return shtname


def is_non_local(is_strategic, is_universal ):

    non_local = False

    if is_strategic == "●" or is_universal == "●":
        non_local = True

    return non_local


def list_chg(temp_list):
    output_list =[]
    for entities in temp_list:
        output_list.append(entities[0])

    return output_list


def find_max(base_data_dict, car_list):

    max_value = 0
    for car_name in car_list:
        if base_data_dict[car_name] is None:
            base_data_dict[car_name] = 0
        if max_value < base_data_dict[car_name]:
            max_value =base_data_dict[car_name]

    #print("max value is " +str(max_value))
    return max_value


def cal_volume(volume_dict, car_list):

    volume = 0

    for car_name in car_list :
        volume = volume + volume_dict[car_name]

    #print("cal volume is "+str(volume))
    return volume


def cal_allocation(conn_string, config_type, base_data_dict):

    conn = pyodbc.connect(conn_string)
    cursor = conn.cursor()

    selectQuery = "select sum(volume) from Volume_old"
    cursor.execute(selectQuery)
    q_list = list(cursor.fetchall())
    volume_total = q_list[0][0]
    car_volume_dict={}

    # print("total volume = " + str(volume_total))

    for i in range (0, 21):
        selectQuery = "select sum(volume) from volume_old where ID='"+car_List[i]+"'"
        cursor.execute(selectQuery)
        q_list = list(cursor.fetchall())
        car_volume_dict[car_List[i]]= q_list[0][0]

    # print(car_volume_dict)

    if config_type == "통합":
        lead_car_value = find_max(base_data_dict, car_List)
        for car_name in base_data_dict.keys():
            base_data_dict[car_name] = lead_car_value*car_volume_dict[car_name]/volume_total

    elif config_type == "차급":

        selectQuery = "select ID from VehicleData where SEG='B' or SEG='C'"
        cursor.execute(selectQuery)
        c_seg_list = list_chg(cursor.fetchall())

        selectQuery = "select ID from VehicleData where SEG='D'"
        cursor.execute(selectQuery)
        d_seg_list = list_chg(cursor.fetchall())

        selectQuery = "select ID from VehicleData where SEG='E' or SEG='E+' "
        cursor.execute(selectQuery)
        e_seg_list = list_chg(cursor.fetchall())

        c_lead_car_value = find_max(base_data_dict, c_seg_list)
        d_lead_car_value = find_max(base_data_dict, d_seg_list)
        e_lead_car_value = find_max(base_data_dict, e_seg_list)

        c_seg_volume = cal_volume(car_volume_dict, c_seg_list)
        d_seg_volume = cal_volume(car_volume_dict, d_seg_list)
        e_seg_volume = cal_volume(car_volume_dict, e_seg_list)

        print("c_lead_car_value : " + str(c_lead_car_value) + "/d_lead_car_value : " + str(d_lead_car_value) + "/e_lead_car_value : " + str(e_lead_car_value))
        print("c_seg_volume : " +str(c_seg_volume) +"/d_seg_volume : "+str(d_seg_volume)+"/e_seg_volume : "+str(e_seg_volume))

        for c_seg_car_name in c_seg_list:
            base_data_dict[c_seg_car_name] = c_lead_car_value * car_volume_dict[c_seg_car_name] / c_seg_volume

        for d_seg_car_name in d_seg_list:
            base_data_dict[d_seg_car_name] = d_lead_car_value * car_volume_dict[d_seg_car_name] / d_seg_volume

        for e_seg_car_name in e_seg_list:
            base_data_dict[e_seg_car_name] = e_lead_car_value * car_volume_dict[e_seg_car_name]/ e_seg_volume

    elif config_type == "브랜드":

        selectQuery = "select ID from VehicleData where Brand='Hyundai' or Brand='Kia' "
        cursor.execute(selectQuery)
        hk_brand_list = list_chg(cursor.fetchall())

        selectQuery = "select ID from VehicleData where Brand='Genesis'"
        cursor.execute(selectQuery)
        g_brand_list = list_chg(cursor.fetchall())

        hk_lead_car_value = find_max(base_data_dict, hk_brand_list)
        g_lead_car_value = find_max(base_data_dict, g_brand_list)

        hk_brand_volume = cal_volume(car_volume_dict, hk_brand_list)
        g_brand_volume = cal_volume(car_volume_dict, g_brand_list)

        for h_brand_car_name in hk_brand_list:
            base_data_dict[h_brand_car_name] = hk_lead_car_value * car_volume_dict[h_brand_car_name] / hk_brand_volume

        for g_brand_car_name in g_brand_list:
            base_data_dict[g_brand_car_name] = g_lead_car_value * car_volume_dict[g_brand_car_name] / g_brand_volume

    elif config_type == "바디":

        selectQuery = "select ID from VehicleData where BT='Hatchback' or BT='Sedan'"
        cursor.execute(selectQuery)
        sedan_list = list_chg(cursor.fetchall())

        selectQuery = "select ID from VehicleData where BT='CUV' or BT='SUV'"
        cursor.execute(selectQuery)
        suv_list = list_chg(cursor.fetchall())

        sedan_lead_car_value = find_max(base_data_dict, sedan_list)
        suv_lead_car_value = find_max(base_data_dict, suv_list)

        sedan_volume = cal_volume(car_volume_dict, sedan_list)
        suv_volume = cal_volume(car_volume_dict, suv_list)

        print("sedan_lead_car_value : " + str(sedan_lead_car_value) + "/suv_lead_car_value : " + str(suv_lead_car_value))
        print("sedan_volume : " +str(sedan_volume) +"/suv_volume : "+str(suv_volume))

        for sedan_car_name in sedan_list:
            base_data_dict[sedan_car_name] = sedan_lead_car_value * car_volume_dict[sedan_car_name] / sedan_volume

        for suv_car_name in suv_list:
            base_data_dict[suv_car_name] = suv_lead_car_value * car_volume_dict[suv_car_name] / suv_volume

    return base_data_dict


def car_no_to_car_name(car_no):

    car_no =int(str.replace(car_no,"ev",""))
    car_name = "VE00"
    if car_no <10 :
        car_name = car_name +"0"+str(car_no)
    if 10 <= car_no < 100:
        car_name = car_name + str(car_no)

    return car_name


def btm_up_chg(file_name):

    targetBk = openpyxl.load_workbook(file_name, data_only=True)
    targetSht = targetBk["BTU"]
    col_location = 106

    while col_location < 147:

        car_name = targetSht.cell(row=4, column=col_location).value

        for i in range(7, 398):

            non_local = is_non_local(targetSht.cell(row=i, column=9).value,targetSht.cell(row=i, column=10).value)

            #print(str(is_strategic) +" and "+str(is_universal))

            base_value = targetSht.cell(row=i, column=col_location).value
            app_value = targetSht.cell(row=i, column=col_location+1).value

            #print(str(i) + " : " + str(is_allocated)+"base_value = "+str(base_value)+",app_value = "+str(app_value))

            if base_value is None:
                None

            elif base_value is not None and app_value is not None:
                None

            elif base_value is not None and app_value is None and non_local is True:
                targetSht.cell(row=i, column=col_location).value = base_value*base_rate
                targetSht.cell(row=i, column=col_location+1).value = base_value * (1-base_rate)

        col_location = col_location + 2

    for i in range(7,398):

        non_local = is_non_local(targetSht.cell(row=i, column=9).value, targetSht.cell(row=i, column=10).value)

        col_location = 106
        config_type = targetSht.cell(row=i,column=104).value
        base_data_dict={}

        if non_local is True :

            while col_location < 147:
                base_data_dict[targetSht.cell(row=4,column=col_location).value]=targetSht.cell(row=i,column=col_location).value
                col_location = col_location + 2

            #print(config_type +" : "+str(base_data_dict))
            base_data_dict = cal_allocation(conn_String,config_type,base_data_dict)
            #print(base_data_dict)
            col_location = 106

            while col_location < 147:

                targetSht.cell(row=i,column=col_location).value=base_data_dict[targetSht.cell(row=4, column=col_location).value]
                if config_type == "X":
                    targetSht.cell(row=i, column=col_location).value = 0
                    targetSht.cell(row=i, column=col_location+1).value = 0
                col_location = col_location + 2

    targetBk.save(file_name)
    targetBk.close()


def check_china_car(car_name, china_car_list):

    is_china = False

    for china_car_name in china_car_list:
        if car_name == china_car_name:
            is_china = True

    return is_china


def btm_up_datacube(file_name,data_file_name) :

    dataBk = openpyxl.load_workbook(data_file_name,read_only=False, keep_vba=True)
    dataSht = dataBk["Data"]
    dataSht.delete_rows(2,  dataSht.max_row + 1)
    data_seq = 2

    targetBk = openpyxl.load_workbook(file_name, data_only=True)
    targetSht = targetBk["BTU"]
    col_location = 106

    while col_location < 147:

        car_name = targetSht.cell(row=4, column=col_location).value
        is_china = check_china_car(car_name,china_car_List)

        for i in range(7, 398):

            non_local = is_non_local(targetSht.cell(row=i, column=9).value, targetSht.cell(row=i, column=10).value)

            base_value = targetSht.cell(row=i, column=col_location).value

            if base_value is None:
                base_value = 0

            app_value = targetSht.cell(row=i, column=col_location+1).value

            if app_value is None:
                app_value = 0
                
            module_name = targetSht.cell(row=i, column=2).value

            if module_name is not None and is_china is False:

                if non_local is True :
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI042", base_value)
                    dataSht = cell_input(dataSht, data_seq+1, car_name, module_name, "NewPAI041", base_value*labor_rate)
                    dataSht = cell_input(dataSht, data_seq+2, car_name, module_name, "NewPAI030", app_value)
                    dataSht = cell_input(dataSht, data_seq+3, car_name, module_name, "NewPAI029", app_value*labor_rate)
                    data_seq = data_seq + 4
                else:
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI028", base_value)
                    dataSht = cell_input(dataSht, data_seq+1, car_name, module_name, "NewPAI026", base_value*labor_rate)
                    data_seq = data_seq + 2

            elif module_name is not None and is_china is True:

                if non_local is True :
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI042", base_value)
                    dataSht = cell_input(dataSht, data_seq+1, car_name, module_name, "NewPAI041", base_value*labor_rate)
                    dataSht = cell_input(dataSht, data_seq+2, car_name, module_name, "NewPAI030", app_value)
                    dataSht = cell_input(dataSht, data_seq+3, car_name, module_name, "NewPAI029", app_value*labor_rate)
                    data_seq = data_seq + 4
                else:
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI028", app_value)
                    dataSht = cell_input(dataSht, data_seq+1, car_name, module_name, "NewPAI026", app_value*labor_rate)
                    data_seq = data_seq + 2

        col_location = col_location +2

    targetBk.close()
    dataBk.save(data_file_name)
    dataBk.close()

#btm_up_chg(file_Name)
btm_up_datacube(file_Name,data_File_Name)