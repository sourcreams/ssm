import openpyxl
import pyodbc
import time

conn_String = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\6408284\Desktop\새 폴더 (2)\moduleBook.accdb;'

file_Name = r"C:\Users\6408284\Desktop\원단위 2 - 12.10 이준우.xlsx"
output_file_Name = r"C:\Users\6408284\Desktop\원단위 2 - 12.10 이준우.xlsx"

sht_Name = "의장_RS4 JW JK NE GL3 CN7 CV SG2"

sht_List = ["BR_YB","BR_NE","BR_OE","BR_OS EV","BR_DL3","BR_JK EV","BR_DL3c","BR_DN8c","BR_TMc"]
sht_china_List = ["BR_DL3c","BR_DN8c","BR_TMc"]

sht_Ref_List = ["JX1","IK","MQ4","BD","TL","DN8","DH"]
rd_column_index_dict = {"TL":4,"JX1":7,"DN8":10,"MQ4":13,"IK":16,"BD":19,"DH":22}

# 시작의장BD/ 시작의장TL,DN8,DH / 시작의장MQ4/ 시작의장IK/ 시작의장JX(전산)
row_index = 9  #  입력 위치
col_index = 6  #  출력 위치
upg_location = 6  # UPG 위치
name_location = 8
base_ratio = 0.65

def name_to_upg(file_Name, sht_List, row_index, col_index, name_location):

    arc_file_Name = r"C:\Users\6408284\Desktop\새 폴더 (2)\ARCHITECTURE.xlsx"

    targetBk = openpyxl.load_workbook(file_Name)
    arcBk = openpyxl.load_workbook(arc_file_Name)
    arcSht = arcBk["ReplaceTable"]

    for sht_name in sht_List:
        targetSht = targetBk[sht_name]

        for i in range(row_index, targetSht.max_row+1):
            for j in range(2, arcSht.max_row+1):
                #print(targetSht.cell(row = i, column=name_location).value)
                if targetSht.cell(row = i, column=name_location).value == arcSht.cell(row=j, column = 5).value :

                    targetSht.cell(row=i, column=col_index).value = arcSht.cell(row=j, column=6).value

    arcBk.close()
    targetBk.save(file_Name)
    targetBk.close()


name_to_upg(file_Name, sht_china_List, row_index, col_index, name_location)